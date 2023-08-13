import pandas as pd
import glob
import openpyxl
from fpdf import FPDF
from pathlib import Path

filepaths = glob.glob("invoices/*.xlsx")

for filepath in filepaths:

    pdf = FPDF(orientation="P", unit="mm", format="A4")
    pdf.add_page()

    filename = Path(filepath).stem
    invoice_no, date = filename.split("-")

    pdf.set_font(family="Times", style="B", size=16)
    pdf.cell(w=50, h=8, txt=f"Invoice No: {invoice_no}", ln=1)

    pdf.set_font(family="Times", style="B", size=16)
    pdf.cell(w=50, h=8, txt=f"Date: {date}", ln=1)

    df = pd.read_excel(filepath, sheet_name="Sheet 1")

    # Add a Header
    columns = df.columns
    columns = [item.replace("_", " ").title() for item in columns]
    pdf.set_font(family="Times", size=10, style="B")
    pdf.set_text_color(80, 80, 80)
    pdf.cell(w=30, h=8, txt=columns[0], border=1)
    pdf.cell(w=70, h=8, txt=columns[1], border=1)
    pdf.cell(w=30, h=8, txt=columns[2], border=1)
    pdf.cell(w=30, h=8, txt=columns[3], border=1)
    pdf.cell(w=30, h=8, txt=columns[4], border=1, ln=1)

    # Added rows to the Table
    for index, row in df.iterrows():
        pdf.set_font(family="Times", size=10)
        pdf.set_text_color(80, 80, 80)
        pdf.cell(w=30, h=8, txt=str(row["product_id"]), border=1)
        pdf.cell(w=70, h=8, txt=str(row["product_name"]), border=1)
        pdf.cell(w=30, h=8, txt=str(row["amount_purchased"]), border=1)
        pdf.cell(w=30, h=8, txt=str(row["price_per_unit"]), border=1)
        pdf.cell(w=30, h=8, txt=str(row["total_price"]), border=1, ln=1)

    # Added Total Sum
    total_sum = df["total_price"].sum()
    pdf.set_font(family="Times", size=10)
    pdf.set_text_color(80, 80, 80)
    pdf.cell(w=30, h=8, txt="", border=1)
    pdf.cell(w=70, h=8, txt="", border=1)
    pdf.cell(w=30, h=8, txt="", border=1)
    pdf.cell(w=30, h=8, txt="", border=1)
    pdf.cell(w=30, h=8, txt=str(total_sum), border=1, ln=1)

    # Added Total Sum Header
    pdf.set_font(family="Times", size=10, style="B")
    pdf.cell(w=30, h=8, txt=f"The Total Price is {total_sum} ", ln=1)

    # Added Footer
    pdf.set_font(family="Times", size=10, style="I")
    pdf.cell(w=30, h=8, txt="Tron Technologies", border=0)
    pdf.image("Tron.png", w=8)

    # Create a PDFs Directory in the Root Folder then proceed
    pdf.output(f"PDFs/{filename}.pdf")
